import pandas as pd
import pathlib
from openpyxl import load_workbook

ROOT = pathlib.Path('.')

# Test files
DEV_LOG = ROOT / '2026 SPRING FALL Development Sample Log_sensitive info removed.xlsx'
TEMPLATE = ROOT / 'Spec Template (1).xlsx'
BOM = ROOT / 'WORKING SIMPLIFIED BOM (1).xlsx'

print("🔍 Testing file structure and data format...")
print("=" * 60)

# Test 1: Development Sample Log
print("\n📋 DEVELOPMENT SAMPLE LOG")
print("-" * 30)
try:
    dev = pd.read_excel(DEV_LOG)
    print(f"✅ File loaded successfully")
    print(f"📊 Shape: {dev.shape}")
    print(f"📝 Initial columns: {list(dev.columns)[:5]}...")  # Show first 5 columns
    
    # Check if first row is header
    print(f"\n🔍 First row (potential header): {dev.iloc[0].tolist()[:5]}...")
    
    # Try promoting first row to header
    dev.columns = dev.iloc[0]
    dev = dev.drop(index=0).reset_index(drop=True)
    print(f"📝 After header promotion: {list(dev.columns)[:5]}...")
    
    # Check for required columns
    required_cols = ['Sample Name', 'Season', 'Gender', 'Factory Ref #', 'Sample Order No.', 'Upper', 'Sole']
    missing_cols = [col for col in required_cols if col not in dev.columns]
    
    if missing_cols:
        print(f"⚠️  Missing columns: {missing_cols}")
        print(f"📝 Available columns: {list(dev.columns)}")
    else:
        print(f"✅ All required columns found")
        print(f"📊 Number of samples: {len(dev)}")
        if len(dev) > 0:
            print(f"📝 Sample names: {dev['Sample Name'].head(3).tolist()}...")

except Exception as e:
    print(f"❌ Error: {e}")

# Test 2: BOM File
print("\n🔍 SIMPLIFIED BOM")
print("-" * 30)
try:
    raw_bom = pd.read_excel(BOM, header=None)
    print(f"✅ File loaded successfully")
    print(f"📊 Shape: {raw_bom.shape}")
    print(f"📝 First few entries:")
    for i, entry in enumerate(raw_bom[0].dropna().head(5)):
        print(f"  {i+1}. {entry}")
    
    # Test the splitting logic
    raw_bom = raw_bom.dropna(how='all').rename(columns={0: 'raw'})
    raw_bom[['Part', 'Material']] = raw_bom['raw'].str.split(':', n=1, expand=True)
    raw_bom['Part'] = raw_bom['Part'].str.strip()
    raw_bom['Material'] = raw_bom['Material'].fillna('').str.strip()
    
    materials = [m for m in raw_bom['Material'] if m]
    print(f"📊 Total materials in BOM: {len(materials)}")
    
except Exception as e:
    print(f"❌ Error: {e}")

# Test 3: Template File
print("\n📝 SPEC TEMPLATE")
print("-" * 30)
try:
    wb = load_workbook(TEMPLATE)
    sheet = wb.active
    print(f"✅ File loaded successfully")
    print(f"📊 Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    print(f"📝 Sheet name: {sheet.title}")
    
    # Look for key cells and labels
    print(f"\n🔍 Key cells:")
    key_cells = ['A1', 'B1', 'C2', 'A4', 'E2']
    for cell in key_cells:
        value = sheet[cell].value
        print(f"  {cell}: {value}")
    
    # Look for part labels in column A
    print(f"\n🔍 Looking for part labels in column A:")
    part_labels = []
    for r in range(1, min(sheet.max_row + 1, 30)):  # Check first 30 rows
        cell_value = sheet[f'A{r}'].value
        if cell_value and ':' in str(cell_value):
            part_labels.append((r, str(cell_value)))
    
    if part_labels:
        print(f"  Found {len(part_labels)} potential part labels:")
        for row, label in part_labels[:10]:  # Show first 10
            print(f"    Row {row}: {label}")
    else:
        print(f"  ⚠️  No part labels found with ':' pattern")

except Exception as e:
    print(f"❌ Error: {e}")

print("\n" + "=" * 60)
print("🏁 File structure test complete!") 