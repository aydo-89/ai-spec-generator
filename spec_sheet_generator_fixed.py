import re, difflib, pathlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = pathlib.Path('.')
# Use the smaller test file as mentioned by the user
DEV_LOG = ROOT / 'Development Sample Log_Spec sheet test (1).xlsx'
TEMPLATE = ROOT / 'Spec Template (1).xlsx'
BOM = ROOT / 'WORKING SIMPLIFIED BOM (1).xlsx'
OUTFILE = ROOT / 'AUTO‑GENERATED SPEC SHEETS.xlsx'

# --------------------------------------------------
# 1.  Load & tidy the Development Sample Log
# --------------------------------------------------
print("📋 Loading Development Sample Log...")
dev = pd.read_excel(DEV_LOG)
print(f"Initial shape: {dev.shape}")
print(f"Initial columns: {list(dev.columns)}")

dev.columns = dev.iloc[0]                      # promote first row → header
dev = dev.drop(index=0).reset_index(drop=True) # drop the header row

print(f"After header promotion: {dev.shape}")
print(f"Final columns: {list(dev.columns)}")

# Check actual column names and map them
column_mapping = {
    'Sample Name': 'Sample Name',
    'Season': 'Season',
    'Gender': 'Gender',
    'Sample Order No.': 'Sample Order No.',
    'Factory ref #': 'Factory Ref #',  # Note the lowercase 'r'
    'Upper': 'Upper',
    'Sole (ref # only)': 'Sole',  # Different column name
}

# Display available data
if len(dev) > 0:
    print(f"Sample names: {dev['Sample Name'].tolist()}")
    print(f"\nFirst sample Upper data: {dev['Upper'].iloc[0] if 'Upper' in dev.columns else 'No Upper column'}")
    print(f"First sample Sole data: {dev['Sole (ref # only)'].iloc[0] if 'Sole (ref # only)' in dev.columns else 'No Sole column'}")

# --------------------------------------------------
# 2.  Build a material normaliser from the Simplified BOM
# --------------------------------------------------
print("\n🔍 Loading Simplified BOM...")
raw_bom = pd.read_excel(BOM, header=None)
print(f"BOM shape: {raw_bom.shape}")
print(f"First few BOM entries: {raw_bom.iloc[:10, :].values}")

# The BOM appears to have a different structure - let's examine it more closely
raw_bom_clean = raw_bom.dropna(how='all')
if raw_bom_clean.shape[1] > 1:
    # If there are multiple columns, check the second column for materials
    print("BOM has multiple columns, checking structure...")
    for i in range(min(10, len(raw_bom_clean))):
        print(f"Row {i}: {raw_bom_clean.iloc[i].tolist()}")

# Create a simple material lookup - we'll improve this with AI later
materials = []
for idx, row in raw_bom_clean.iterrows():
    for col in row:
        if pd.notna(col) and str(col).strip() and ':' not in str(col):
            materials.append(str(col).strip())

# Remove duplicates and common non-materials
materials = list(set([m for m in materials if len(m) > 2 and m not in ['Upper', 'Lining', 'Sole']]))
print(f"Extracted {len(materials)} potential materials: {materials[:10]}...")

# Create lookup dictionaries
exact_lookup = {m.lower(): m for m in materials}
all_materials = list(exact_lookup.keys())

def standardise(text: str) -> str:
    """Return canonical material name from the BOM (exact → fuzzy)."""
    if not isinstance(text, str):
        return text
    txt = text.strip().lower()
    if txt in exact_lookup:
        return exact_lookup[txt]

    # fuzzy fallback
    match = difflib.get_close_matches(txt, all_materials, n=1, cutoff=0.70)
    return exact_lookup[match[0]] if match else text   # leave unchanged if no hit

# --------------------------------------------------
# 3.  Enhanced helper to parse material descriptions
# --------------------------------------------------
def parse_complex_materials(text: str) -> dict:
    """
    Parse complex material descriptions like your example:
    'Upper: W1063 Minnetonka Brown Cow Suede - lining: Microfiber, Color: Brown...'
    """
    if not isinstance(text, str):
        return {}
    
    materials = {}
    text = text.replace('\n', ' ')  # Normalize line breaks
    
    # Look for patterns like "Part:" or "Part/Other:"
    patterns = [
        r'(Upper[^:]*?):\s*([^-]+?)(?:\s*-|$)',
        r'(Lining[^:]*?):\s*([^,]+?)(?:\s*,|$)',
        r'(Sole[^:]*?):\s*([^,]+?)(?:\s*,|$)',
        r'(Midsole[^:]*?):\s*([^,]+?)(?:\s*,|$)',
        r'(Outsole[^:]*?):\s*([^,]+?)(?:\s*,|$)',
        r'(Footbed[^:]*?):\s*([^,]+?)(?:\s*,|$)',
        r'(Insole[^:]*?):\s*([^,]+?)(?:\s*,|$)',
    ]
    
    for pattern in patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE)
        for match in matches:
            part = match.group(1).strip()
            material = match.group(2).strip()
            
            # Clean up the part name
            part = re.sub(r'[/\s]+', ' ', part.strip())
            
            # Clean up the material (remove color info, etc.)
            material = re.split(r'[,\-]\s*(?:Color|Colour)', material)[0].strip()
            
            if part and material and len(material) > 2:
                materials[part] = standardise(material)
    
    print(f"  📝 Parsed materials: {materials}")
    return materials

# --------------------------------------------------
# 4.  Open the blank template and examine structure
# --------------------------------------------------
print(f"\n📝 Loading template from {TEMPLATE}...")
wb_master = load_workbook(TEMPLATE)
base_sheet: Worksheet = wb_master.active

print(f"Template loaded. Active sheet: {base_sheet.title}")
print(f"Template dimensions: {base_sheet.max_row} rows x {base_sheet.max_column} columns")

# Let's examine the template structure more carefully
print("\n🔍 Template structure:")
for r in range(1, min(32, base_sheet.max_row + 1)):  # Check first 31 rows
    a_cell = base_sheet[f'A{r}'].value
    b_cell = base_sheet[f'B{r}'].value
    if a_cell and str(a_cell).strip():
        print(f"  Row {r}: A{r}='{a_cell}' | B{r}='{b_cell}'")

processed_count = 0
for idx, row in dev.iterrows():
    try:
        sample_name = row['Sample Name']
        print(f"\n🔧 Processing sample: {sample_name}")
        
        sheet = wb_master.copy_worksheet(base_sheet)
        # Clean sheet name
        clean_name = re.sub(r'[\\/*?[\]:]+', '_', str(sample_name))
        sheet.title = (clean_name[:28] + '...') if len(clean_name) > 31 else clean_name

        # Fill metadata - adjust cell positions based on actual template
        season_gender = f"{row.get('Season', '')}, {row.get('Gender', '')}"
        sheet['B1'] = season_gender  # Based on template analysis
        
        # Try different cells for different data points
        if 'Factory ref #' in row:
            sheet['A4'] = row['Factory ref #']
        sheet['C2'] = sample_name
        if 'Sample Order No.' in row:
            sheet['E2'] = row['Sample Order No.']

        # Parse materials from the complex text
        upper_materials = parse_complex_materials(row.get('Upper', ''))
        
        # Handle the sole column (which is named differently)
        sole_text = row.get('Sole (ref # only)', '') or row.get('Sole', '')
        sole_materials = parse_complex_materials(sole_text)

        # Try to fill template based on what we find
        all_materials = {**upper_materials, **sole_materials}
        
        def find_template_row(part_name: str) -> int:
            """Find where to put a material in the template"""
            # Look for the part name in column A
            for r in range(1, sheet.max_row + 1):
                cell_value = sheet[f'A{r}'].value
                if cell_value and part_name.lower() in str(cell_value).lower():
                    return r
            return None

        materials_filled = 0
        for part, material in all_materials.items():
            template_row = find_template_row(part)
            if template_row:
                sheet[f'B{template_row}'] = material
                materials_filled += 1
                print(f"  ✅ {part}: {material} → Row {template_row}")
            else:
                print(f"  ⚠️  Couldn't find template location for {part}")

        print(f"  📊 Filled {materials_filled} materials")
        processed_count += 1
        
    except Exception as e:
        print(f"  ❌ Error processing {sample_name}: {e}")
        import traceback
        traceback.print_exc()
        continue

# --------------------------------------------------
# 5.  Save
# --------------------------------------------------
print(f"\n💾 Saving workbook...")
wb_master.remove(base_sheet)
wb_master.save(OUTFILE)
print(f'✅ Created {OUTFILE}')
print(f'📊 Successfully processed {processed_count} samples out of {len(dev)} total') 