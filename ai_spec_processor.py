import re
import json
import difflib
import pathlib
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Optional, Tuple
import openai
from dataclasses import dataclass
import logging
from io import BytesIO
from fuzzywuzzy import fuzz
import openpyxl.utils

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class MaterialMatch:
    original: str
    standardized: str
    confidence: float
    method: str  # 'exact', 'fuzzy', 'ai', 'no_match'
    reasoning: Optional[str] = None

@dataclass
class ProcessingResult:
    success: bool
    samples_processed: int
    total_samples: int
    matches_by_method: Dict[str, int]
    errors: List[str]
    output_file: Optional[bytes] = None

class AISpecProcessor:
    def __init__(self, api_key: str):
        """Initialize the AI-enhanced spec processor"""
        self.client = openai.OpenAI(api_key=api_key)
        self.ai_cache = {}  # Cache AI results to avoid repeated API calls
        self.materials = []
        self.exact_lookup = {}
        self.material_codes = {}  # Maps codes to full descriptions
        
        # Template field mappings based on analysis
        self.template_mapping = {
            'season_gender': ('B', 1),
            'sample_name': ('G', 3),
            'color_name': ('L', 3),
            'factory_ref': ('B', 5),
            'sample_number': ('B', 26),
            'upper': ('H', 6),
            'trim': ('H', 8),
            'lining': ('H', 9),
            'sock_topcover': ('H', 10),
            'sock_label': ('H', 11),
            'insole': ('H', 12),
            'midsole': ('H', 13),
            'outsole': ('H', 14),
            'outsole_treatment': ('H', 15),
            'detail_stitching': ('H', 16),
            'reg_stitching': ('H', 17),
            'hardware': ('H', 18),
            'other': ('H', 19),
            'notes': ('B', 27)  # default, may be overwritten by _refresh_template_mapping
        }
        
    def load_bom(self, bom_file: BytesIO) -> bool:
        """Load and process the BOM file with proper material extraction"""
        try:
            logger.info("Loading BOM file...")
            raw_bom = pd.read_excel(bom_file)
            logger.info(f"BOM shape: {raw_bom.shape}")
            logger.info(f"BOM columns: {list(raw_bom.columns)}")
            
            materials = []
            material_codes = {}
            
            # Extract all non-null string values from BOM
            for idx, row in raw_bom.iterrows():
                for col_val in row:
                    if pd.notna(col_val) and isinstance(col_val, str) and len(col_val.strip()) > 2:
                        material = col_val.strip()
                        materials.append(material)
                        
                        # Extract material codes (patterns like W1063, MBA2, VP-052, etc.)
                        code_matches = re.findall(r'\b([A-Z]+\d+(?:-\d+)?)\b', material)
                        for code in code_matches:
                            material_codes[code] = material
                            material_codes[code.lower()] = material
            
            # Clean and deduplicate
            self.materials = list(set(materials))
            self.exact_lookup = {m.lower(): m for m in self.materials}
            self.material_codes = material_codes
            
            logger.info(f"Loaded {len(self.materials)} materials from BOM")
            logger.info(f"Extracted {len(self.material_codes)} material codes")
            logger.info(f"Sample materials: {self.materials[:5]}")
            logger.info(f"Sample codes: {list(self.material_codes.keys())[:10]}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error loading BOM: {e}")
            return False
    
    def parse_vendor_materials(self, upper_text: str, sole_text: str, model: str = "gpt-4.1-mini") -> Dict[str, str]:
        """AI-enhanced parsing of vendor material descriptions with intelligent categorization"""
        if not isinstance(upper_text, str):
            upper_text = ""
        if not isinstance(sole_text, str):
            sole_text = ""
            
        logger.info(f"Parsing materials - Upper: '{upper_text}', Sole: '{sole_text}'")
            
        try:
            # Combine all material text for AI analysis
            combined_text = f"Upper materials: {upper_text}\nSole materials: {sole_text}"
            
            # Build category section dynamically from template keys
            cat_lines = []
            key_to_label = {
                'upper': 'Upper: Main upper materials',
                'trim': 'Trim: Decorative elements, laces, bindings',
                'lining': 'Lining: Interior lining',
                'sock_topcover': 'Sock (topcover): Footbed / sock liner',
                'sock_label': 'Sock Label: Sock labels or prints',
                'insole': 'Insole: Insole materials',
                'midsole': 'Midsole: Midsole materials',
                'outsole': 'Outsole: Outsole / bottom',
                'outsole_treatment': 'Outsole Treatment: Special outsole treatments',
                'detail_stitching': 'Detail Stitching: Decorative stitching',
                'reg_stitching': 'Reg Stitching: Regular stitching',
                'hardware': 'Hardware: Buckles, eyelets, metal components',
                'other': 'Other: Anything that does not fit above'
            }
            for k in self.template_mapping.keys():
                if k in key_to_label:
                    cat_lines.append(f"- {key_to_label[k]}")
            categories_block = "\n".join(cat_lines)

            prompt = f"""You are a footwear material categorization expert. Parse these vendor material descriptions and categorize them into the correct spec sheet fields.

VENDOR MATERIALS:\n{combined_text}

SPEC SHEET CATEGORIES:\n{categories_block}

VENDOR TERMINOLOGY VARIATIONS:
- "Ball stitching", "Small stitching" → Detail Stitching or Reg Stitching
- "Pile", "Faux fur" → Upper or Lining
- "TPR", "EVA", "Rubber" → Outsole
- "Microfiber", "Berber" → Lining
- Material codes like "VP-052", "W211" → Extract the material name

CRITICAL REQUIREMENTS:
1. NOTHING can be omitted - every piece of information must be categorized
2. Preserve ALL original details and descriptions exactly as written
3. If unsure about a category, put it in "Other" with the original label
4. Keep material codes, numbers, colors, and descriptions intact
5. Handle multi-part numbers like "M63 - 360" or "90 360-81343" as single units

RULES:
1. Parse colon-delimited entries (e.g., "Upper: Brown Suede")
2. Handle vendor-specific terminology intelligently
3. If unsure about category, put in "Other" with format "Label: Material"
4. Preserve complete descriptions: "HB2085 tan recycled faux fur" stays complete
5. Keep color information: "Color: Brown matching upper" stays complete
6. Maintain material codes: "VP-052, Natural, textile covered" stays complete
7. Never leave materials uncategorized

Return ONLY valid JSON with ALL information categorized:
{{
    "Upper": "complete material description or empty string",
    "Trim": "complete material description or empty string",
    "Lining": "complete material description or empty string",
    "Sock (topcover)": "complete material description or empty string",
    "Sock Label": "complete material description or empty string",
    "Insole": "complete material description or empty string",
    "Midsole": "complete material description or empty string",
    "Outsole": "complete material description or empty string",
    "Outsole Treatment": "complete material description or empty string",
    "Detail Stitching": "complete material description or empty string",
    "Reg Stitching": "complete material description or empty string",
    "Hardware": "complete material description or empty string",
    "Other": "Label: Complete description (for any unclear items)"
}}"""

            response = self.client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=800
            )
            
            result_text = response.choices[0].message.content.strip()
            if result_text.startswith('```json'):
                result_text = result_text.replace('```json', '').replace('```', '').strip()
            
            parsed_materials = json.loads(result_text)
            logger.info(f"AI parsed materials: {parsed_materials}")
            
            # Clean up empty values
            cleaned_materials = {k: v for k, v in parsed_materials.items() if v and v.strip()}
            logger.info(f"Cleaned materials: {cleaned_materials}")
            return cleaned_materials
            
        except Exception as e:
            logger.warning(f"AI parsing error: {e}")
            # Fallback to regex parsing
            fallback_result = self._regex_parse_materials(upper_text, sole_text)
            logger.info(f"Fallback parsing result: {fallback_result}")
            return fallback_result
    
    def _regex_parse_materials(self, upper_text: str, sole_text: str) -> Dict[str, str]:
        """Fallback regex parsing for vendor materials"""
        materials = {}
        
        logger.info(f"Using fallback regex parsing for Upper: '{upper_text}', Sole: '{sole_text}'")
        
        # Enhanced patterns for common formats
        patterns = [
            (r'upper[^:]*?:\s*([^,\n]+)', 'Upper'),
            (r'lining[^:]*?:\s*([^,\n]+)', 'Lining'),
            (r'sole[^:]*?:\s*([^,\n]+)', 'Outsole'),
            (r'outsole[^:]*?:\s*([^,\n]+)', 'Outsole'),
            (r'midsole[^:]*?:\s*([^,\n]+)', 'Midsole'),
            (r'trim[^:]*?:\s*([^,\n]+)', 'Trim'),
            (r'hardware[^:]*?:\s*([^,\n]+)', 'Hardware'),
        ]
        
        combined_text = f"{upper_text} {sole_text}"
        
        for pattern, category in patterns:
            matches = re.finditer(pattern, combined_text, re.IGNORECASE)
            for match in matches:
                material = match.group(1).strip()
                if material and len(material) > 2:
                    materials[category] = material
                    logger.info(f"Regex matched {category}: {material}")
                    break
        
        # If no patterns match, put raw materials in appropriate categories
        if not materials:
            logger.info("No regex patterns matched, using raw categorization")
            if upper_text.strip():
                # Try to parse comma-separated items in upper text
                upper_parts = [part.strip() for part in upper_text.split(',') if part.strip()]
                for i, part in enumerate(upper_parts):
                    if i == 0:
                        materials['Upper'] = part
                    else:
                        # Additional upper materials go to Other
                        if 'Other' not in materials:
                            materials['Other'] = f"Upper material: {part}"
                        else:
                            materials['Other'] += f"; Upper material: {part}"
                logger.info(f"Raw Upper: {materials.get('Upper', '')}")
            
            if sole_text.strip():
                # Try to parse comma-separated items in sole text
                sole_parts = [part.strip() for part in sole_text.split(',') if part.strip()]
                for i, part in enumerate(sole_parts):
                    if i == 0:
                        materials['Outsole'] = part
                    else:
                        # Additional sole materials go to Other
                        if 'Other' not in materials:
                            materials['Other'] = f"Sole material: {part}"
                        else:
                            materials['Other'] += f"; Sole material: {part}"
                logger.info(f"Raw Outsole: {materials.get('Outsole', '')}")
        
        return materials
    
    def standardize_material(self, raw_material: str, model: str = "gpt-4.1-mini") -> MaterialMatch:
        """Enhanced material standardization with BOM cross-reference and 90% confidence threshold"""
        if not isinstance(raw_material, str) or not raw_material.strip():
            return MaterialMatch(raw_material, raw_material, 0.0, 'no_match')
        
        clean_material = raw_material.strip()
        logger.info(f"Enhanced BOM matching for: '{clean_material}'")
        
        # 1. Check for exact material code matches first
        material_codes = re.findall(r'\b([A-Z]+\d+(?:-\d+)?)\b', clean_material)
        for code in material_codes:
            if code in self.material_codes:
                logger.info(f"Found exact code match: {code} -> {self.material_codes[code]}")
                return MaterialMatch(
                    original=raw_material,
                    standardized=self.material_codes[code],
                    confidence=1.0,
                    method='exact',
                    reasoning=f"Exact code match for {code}"
                )
            elif code.lower() in self.material_codes:
                logger.info(f"Found case-insensitive code match: {code} -> {self.material_codes[code.lower()]}")
                return MaterialMatch(
                    original=raw_material,
                    standardized=self.material_codes[code.lower()],
                    confidence=1.0,
                    method='exact',
                    reasoning=f"Exact code match for {code}"
                )
        
        # 2. Try exact string match
        material_lower = clean_material.lower()
        if material_lower in self.exact_lookup:
            logger.info(f"Found exact string match: {self.exact_lookup[material_lower]}")
            return MaterialMatch(
                original=raw_material,
                standardized=self.exact_lookup[material_lower],
                confidence=1.0,
                method='exact',
                reasoning="Exact string match"
            )
        
        # 3. Try high-confidence fuzzy matching (90%+)
        best_match = None
        best_ratio = 0
        
        for bom_material in self.materials:
            ratio = fuzz.token_set_ratio(material_lower, bom_material.lower()) / 100.0
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = bom_material
        
        # Require at least one token overlap to prevent random matches like "camel berber"
        token_overlap = bool(set(re.findall(r"[a-z0-9']+", material_lower)) & set(re.findall(r"[a-z0-9']+", best_match.lower()))) if best_match else False

        if best_match and best_ratio >= 0.90 and token_overlap:
            logger.info(f"Found high-confidence fuzzy match ({best_ratio:.2f}): {best_match}")
            return MaterialMatch(
                original=raw_material,
                standardized=best_match,
                confidence=best_ratio,
                method='fuzzy',
                reasoning=f"High-confidence fuzzy match ({best_ratio:.2f})"
            )
        
        # 4. Try AI-enhanced matching for partial descriptions
        ai_result = self._ai_material_match(clean_material, model=model)
        if ai_result and ai_result.confidence >= 0.90:
            logger.info(f"Found AI match: {ai_result.standardized}")
            return ai_result
        
        # 5. Conservative approach - return original if not confident
        logger.info(f"No confident match found, returning original: {clean_material}")
        return MaterialMatch(
            original=raw_material,
            standardized=raw_material,
            confidence=0.0,
            method='no_match',
            reasoning="No confident match found in BOM"
        )
    
    def _ai_material_match(self, material: str, model: str = "gpt-4.1-mini") -> Optional[MaterialMatch]:
        """Use AI to find the best BOM material match for partial descriptions"""
        if material in self.ai_cache:
            return self.ai_cache[material]
        
        try:
            # Create prompt with BOM materials
            bom_sample = self.materials[:50]  # Use more materials for better matching
            bom_list = "\n".join([f"- {mat}" for mat in bom_sample])
            
            prompt = f"""You are a footwear material matching expert. Find the best match for this vendor material description from the BOM.

VENDOR MATERIAL: "{material}"

BOM MATERIALS (first 50):
{bom_list}

MATCHING RULES:
1. Look for exact code matches (e.g., "MBA2" should match "MBA2 color tan")
2. Look for partial matches (e.g., "Brown Suede" → "W1063 Minnetonka Brown Cow Suede")
3. Consider synonyms (e.g., "Pile" → "Faux Fur", "TPR" → "Rubber")
4. Match colors and material types
5. Preserve complete BOM descriptions including codes and full names
6. Only return matches with 90%+ confidence
7. If not 90%+ confident, return null

Return ONLY valid JSON:
{{
    "best_match": "exact BOM material name or null if confidence < 90%",
    "confidence": 0.0-1.0,
    "reasoning": "brief explanation of the match or why no match"
}}"""

            response = self.client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=300
            )
            
            result_text = response.choices[0].message.content.strip()
            if result_text.startswith('```json'):
                result_text = result_text.replace('```json', '').replace('```', '').strip()
            
            result = json.loads(result_text)
            
            if result.get('best_match') and result.get('confidence', 0) >= 0.90:
                # Verify token overlap to avoid unrelated matches
                src_tokens = set(re.findall(r"[a-z0-9']+", material.lower()))
                match_tokens = set(re.findall(r"[a-z0-9']+", result['best_match'].lower()))
                if src_tokens & match_tokens:
                    match = MaterialMatch(
                        original=material,
                        standardized=result['best_match'],
                        confidence=result['confidence'],
                        method='ai',
                        reasoning=result.get('reasoning')
                    )
                    self.ai_cache[material] = match
                    return match
                else:
                    logger.info(f"AI suggested match '{result['best_match']}' rejected due to no token overlap with '{material}'")
            
        except Exception as e:
            logger.warning(f"AI BOM matching error for '{material}': {e}")
        
        return None
    
    def infer_color_name(self, materials_dict: Dict[str, str], model: str = "gpt-4.1-mini") -> str:
        """AI-powered color inference from material descriptions"""
        try:
            # Combine all materials for color analysis
            all_materials = " ".join([f"{k}: {v}" for k, v in materials_dict.items() if v])
            
            prompt = f"""Extract the primary color name from these footwear materials:

MATERIALS: {all_materials}

RULES:
- Return the main/primary color (e.g., "Brown", "Black", "Natural", "Cognac")
- Look in Upper materials first, then other materials
- For multi-color items, pick the dominant color
- Use standard color names (not codes)
- If no clear color, return "Natural"

Return ONLY the color name (one or two words max):"""

            response = self.client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=50
            )
            
            color = response.choices[0].message.content.strip().strip('"')
            return color if color else "Natural"
            
        except Exception as e:
            logger.warning(f"Color inference error: {e}")
            return "Natural"
    
    def validate_materials(self, upper_text: str, sole_text: str, parsed_dict: Dict[str, str], model: str = "gpt-4.1-mini") -> Dict[str, str]:
        """Run an additional AI validation pass to ensure nothing is omitted and everything is categorized correctly.

        The validator compares the original vendor text with the initially-parsed dictionary and returns a corrected
        dictionary that:
        1. Preserves every piece of information found in the vendor text.
        2. Moves any uncategorized items into the appropriate category or into "Other".
        3. Never removes or alters existing information unless it is clearly duplicated.
        """
        try:
            combined_text = f"Upper materials: {upper_text}\nSole materials: {sole_text}"
            draft_json = json.dumps(parsed_dict, ensure_ascii=False)

            prompt = f"""You are validating a footwear materials parsing task. Below is the ORIGINAL vendor text followed by an INITIAL JSON categorization.  
If ANY detail from the vendor text is missing from the JSON or is in the wrong category, produce a corrected JSON with ALL details captured.  
If you are unsure where something belongs, put it in \"Other\" as \"Label: material description\".  
NEVER delete or shorten descriptions.  
Return ONLY valid JSON with the exact same keys that were provided.

ORIGINAL VENDOR TEXT:
{combined_text}

INITIAL JSON:
{draft_json}
"""
            response = self.client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=800
            )
            result_text = response.choices[0].message.content.strip()
            if result_text.startswith('```json'):
                result_text = result_text.replace('```json', '').replace('```', '').strip()
            corrected = json.loads(result_text)
            # Fallback to initial dict if something went wrong
            if isinstance(corrected, dict):
                return {k: v for k, v in corrected.items() if v and str(v).strip()}
        except Exception as e:
            logger.warning(f"Validation AI error: {e}")
        return parsed_dict
    
    def _post_process_materials(self, parsed: Dict[str, str]) -> Dict[str, str]:
        """Heuristic fixes after AI + validation parsing to catch stitching/notes issues"""
        adjusted = parsed.copy()
        # Move any 'stitch' phrases out of generic categories into stitching buckets
        for key in list(adjusted.keys()):
            val = adjusted[key]
            low = val.lower()
            # simple note detection
            if re.search(r"\b(advise|measure|please|ensure|inform|confirm|provide)\b", low):
                adjusted.setdefault('notes', '')
                adjusted['notes'] += (('; ' if adjusted['notes'] else '') + val)
                del adjusted[key]
                continue
            if 'stitch' in low:
                # Decide detail vs reg
                target_key = 'detail_stitching' if 'detail' in low or 'cross' in low or 'ball' in low else 'reg_stitching'
                if target_key not in adjusted or not adjusted[target_key]:
                    adjusted[target_key] = val
                    del adjusted[key]
        # remove duplicates across categories (except other/notes)
        seen = set()
        for k in list(adjusted.keys()):
            v = adjusted[k]
            if v in seen and k not in ('other', 'notes'):
                del adjusted[k]
            else:
                seen.add(v)
        return adjusted
    
    def _qa_review(self, original_text: str, parsed_dict: Dict[str, str], model: str = "gpt-4.1-mini") -> Dict[str, str]:
        """Ask LLM to list any substrings missing from parsed_dict; append them to Other."""
        try:
            prompt = f"""Compare the ORIGINAL vendor text with the JSON below.  Return JSON with two arrays: missing (any phrase that does not appear in the JSON values) and ok if nothing is missing.

ORIGINAL:\n{original_text}
JSON:\n{json.dumps(parsed_dict, ensure_ascii=False)}"""
            resp = self.client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.0,
                max_tokens=200,
                response_format={"type": "json_object"}
            )
            data = json.loads(resp.choices[0].message.content)
            missing = data.get('missing', []) if isinstance(data, dict) else []
            if missing:
                other_val = parsed_dict.get('Other', '')
                for m in missing:
                    other_val += (('; ' if other_val else '') + m)
                parsed_dict['Other'] = other_val
            return parsed_dict
        except Exception as e:
            logger.warning(f"QA review failed: {e}")
            return parsed_dict
    
    def process_spec_sheets(self, dev_log_file: BytesIO, template_file: BytesIO, model: str = "gpt-4.1-mini") -> ProcessingResult:
        """Main processing function with comprehensive fixes based on user feedback"""
        try:
            logger.info("Starting enhanced spec sheet processing...")
            
            # Load development log with correct header handling
            dev = pd.read_excel(dev_log_file)
            logger.info(f"Initial development log shape: {dev.shape}")
            logger.info(f"Initial development log columns: {list(dev.columns)}")
            
            # The file analysis showed the actual data starts at row 1, not row 0.
            # Use the first row as headers if standard headers aren't present.
            if 'Sample Name' not in dev.columns:
                dev.columns = dev.iloc[0]
                dev = dev.drop(index=0).reset_index(drop=True)
                logger.info(f"Used first row as headers. New columns: {list(dev.columns)}")
            
            logger.info(f"Loaded {len(dev)} samples from development log")
            
            # Load template
            wb_master = load_workbook(template_file)
            base_sheet = wb_master.active
            
            self._refresh_template_mapping(base_sheet)
            
            stats = {
                'exact_matches': 0, 'fuzzy_matches': 0, 'ai_matches': 0, 'no_matches': 0
            }
            errors = []
            processed_count = 0
            
            for idx, row in dev.iterrows():
                try:
                    sample_name = str(row.get('Sample Name', f'Sample_{idx}'))
                    logger.info(f"Processing sample: {sample_name}")
                    
                    sheet = wb_master.copy_worksheet(base_sheet)
                    clean_name = re.sub(r'[\\/*?[\]:]+', '_', sample_name)
                    sheet.title = (clean_name[:28] + '...') if len(clean_name) > 31 else clean_name
                    
                    # --- Fill Metadata ---
                    sheet['G3'] = sample_name
                    season = str(row.get('Season', ''))
                    gender = str(row.get('Gender', ''))
                    sheet['B1'] = f"{season}, {gender}".strip(', ')
                    
                    # Improved factory reference retrieval (case-insensitive, flexible column names)
                    factory_ref_columns = [c for c in dev.columns if 'factory' in str(c).lower() and 'ref' in str(c).lower()]
                    factory_ref_val = ''
                    for fr_col in factory_ref_columns:
                        val = str(row.get(fr_col, ''))
                        if val and 'nan' not in val.lower():
                            factory_ref_val = val
                            break
                    if factory_ref_val:
                        sheet['B5'] = factory_ref_val
                    
                    # Sample number (flexible lookup)
                    sample_no_cols = [c for c in dev.columns if 'sample' in str(c).lower() and ('order' in str(c).lower() or 'number' in str(c).lower())]
                    sample_no_val = ''
                    for sn_col in sample_no_cols:
                        val = str(row.get(sn_col, ''))
                        if val and 'nan' not in val.lower():
                            sample_no_val = val
                            break
                    if sample_no_val:
                        sheet['B26'] = sample_no_val
                        
                    # --- Comprehensive Material Parsing ---
                    upper_text = str(row.get('Upper', ''))
                    sole_text = str(row.get('Sole (ref # only)', '') or str(row.get('Sole', '')))
                    
                    parsed_materials = self.parse_vendor_materials(upper_text, sole_text, model=model)
                    # NEW – run validation pass to ensure nothing omitted
                    parsed_materials = self.validate_materials(upper_text, sole_text, parsed_materials, model=model)
                    # Heuristic post-processing
                    parsed_materials = self._post_process_materials(parsed_materials)
                    # QA completeness review
                    parsed_materials = self._qa_review(f"Upper: {upper_text}\nSole: {sole_text}", parsed_materials, model=model)
                    logger.info(f"Parsed materials for {sample_name}: {parsed_materials}")
                    
                    # --- Infer Color ---
                    color_name = self.infer_color_name(parsed_materials, model=model)
                    sheet['L3'] = color_name
                    
                    # --- Process Materials (Upper First) ---
                    processed_materials = {}
                    
                    # Helper function to update stats correctly
                    def update_stats(method: str):
                        if method == 'no_match':
                            stats['no_matches'] += 1
                        else:
                            stats[f'{method}_matches'] += 1
                    
                    # Process Upper first to handle "matching upper" dependencies
                    upper_material_standardized = ""
                    if parsed_materials.get('Upper'):
                        match_result = self.standardize_material(parsed_materials['Upper'], model=model)
                        upper_material_standardized = match_result.standardized
                        processed_materials['upper'] = upper_material_standardized
                        update_stats(match_result.method)
                    
                    # Process remaining materials
                    for category, raw_material in parsed_materials.items():
                        # Normalize category names to match template_mapping keys
                        category_key = category.lower().replace(' ', '_').replace('(', '').replace(')', '')
                        if category_key == 'upper':
                            continue

                        # Handle "matching upper" substitutions
                        if 'matching upper' in raw_material.lower() and upper_material_standardized:
                            final_material = re.sub('matching upper', upper_material_standardized, raw_material, flags=re.IGNORECASE)
                            logger.info(f"Substituted 'matching upper' in {category}: {final_material}")
                        else:
                            match_result = self.standardize_material(raw_material, model=model)
                            final_material = match_result.standardized
                            update_stats(match_result.method)
                        
                        processed_materials[category_key] = final_material

                    # --- Fill Spec Sheet ---
                    for category_key, material_value in processed_materials.items():
                        try:
                            if category_key in self.template_mapping:
                                col, row_num = self.template_mapping[category_key]
                                sheet[f'{col}{row_num}'] = material_value
                                logger.info(f"Filled {col}{row_num} ({category_key}) with: {material_value}")
                            else:
                                # Put uncategorized items in "Other"
                                other_col, other_row = self.template_mapping['other']
                                current_other = sheet[f'{other_col}{other_row}'].value or ""
                                new_other_entry = f"{category_key}: {material_value}"
                                sheet[f'{other_col}{other_row}'] = f"{current_other}; {new_other_entry}" if current_other else new_other_entry
                                logger.info(f"Appended to Other: {new_other_entry}")
                        except Exception as e:
                            logger.warning(f"Error filling spec sheet for {category_key}: {e}")
                            # Try to put in Other as fallback
                            try:
                                other_col, other_row = self.template_mapping['other']
                                current_other = sheet[f'{other_col}{other_row}'].value or ""
                                new_other_entry = f"{category_key}: {material_value}"
                                sheet[f'{other_col}{other_row}'] = f"{current_other}; {new_other_entry}" if current_other else new_other_entry
                                logger.info(f"Fallback - Appended to Other: {new_other_entry}")
                            except Exception as e2:
                                logger.error(f"Failed to add to Other category: {e2}")

                    processed_count += 1
                    
                except Exception as e:
                    error_msg = f"Error processing sample '{sample_name}': {str(e)}"
                    errors.append(error_msg)
                    logger.error(error_msg, exc_info=True)
            
            # Remove original template sheet
            if base_sheet.title in wb_master.sheetnames:
                 wb_master.remove(base_sheet)
            
            # Save to BytesIO
            output_buffer = BytesIO()
            wb_master.save(output_buffer)
            output_buffer.seek(0)
            
            return ProcessingResult(
                success=True,
                samples_processed=processed_count,
                total_samples=len(dev),
                matches_by_method=stats,
                errors=errors,
                output_file=output_buffer.getvalue()
            )
            
        except Exception as e:
            logger.error(f"Top-level processing failed: {e}", exc_info=True)
            return ProcessingResult(success=False, samples_processed=0, total_samples=0, matches_by_method={}, errors=[str(e)])

    def _refresh_template_mapping(self, sheet: Worksheet):
        """Scan the template sheet at runtime and update self.template_mapping based on placeholder text.
        This makes the pipeline resilient to future template edits without needing code changes."""
        try:
            dynamic_map = {}
            for row in range(1, 30):
                for col in [2, 8, 12]:  # B, H, L columns likely to contain placeholders
                    cell_val = str(sheet.cell(row=row, column=col).value or '').lower()
                    if not cell_val:
                        continue
                    # Simple keyword→category heuristics
                    key = None
                    if 'upper' in cell_val:
                        key = 'upper'
                    elif 'trim' in cell_val:
                        key = 'trim'
                    elif 'lining' in cell_val and 'sock' not in cell_val:
                        key = 'lining'
                    elif 'sock' in cell_val and 'label' in cell_val:
                        key = 'sock_label'
                    elif 'sock' in cell_val:
                        key = 'sock_topcover'
                    elif 'insole' in cell_val:
                        key = 'insole'
                    elif 'midsole' in cell_val:
                        key = 'midsole'
                    elif 'outsole treatment' in cell_val or 'outsole_treatment' in cell_val:
                        key = 'outsole_treatment'
                    elif 'outsole' in cell_val:
                        key = 'outsole'
                    elif 'detail stitch' in cell_val or 'detail_stitch' in cell_val:
                        key = 'detail_stitching'
                    elif 'reg stitch' in cell_val or 'regular stitch' in cell_val:
                        key = 'reg_stitching'
                    elif 'hardware' in cell_val:
                        key = 'hardware'
                    elif 'other' in cell_val:
                        key = 'other'
                    elif 'note' in cell_val:
                        key = 'notes'
                    # metadata fields
                    elif 'season' in cell_val and 'gender' in cell_val:
                        key = 'season_gender'
                    elif 'sample name' in cell_val:
                        key = 'sample_name'
                    elif 'factory ref' in cell_val:
                        key = 'factory_ref'
                    elif 'sample number' in cell_val or 'sample no' in cell_val:
                        key = 'sample_number'
                    elif 'color' in cell_val and 'name' in cell_val:
                        key = 'color_name'
                    if key:
                        dynamic_map[key] = (openpyxl.utils.get_column_letter(col), row)
            # Merge with existing map – fallback to defaults if not discovered
            self.template_mapping.update(dynamic_map)
            logger.info(f"Template mapping auto-refreshed: {self.template_mapping}")
        except Exception as e:
            logger.warning(f"Template mapping refresh failed, using defaults: {e}") 