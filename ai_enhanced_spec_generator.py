import re, difflib, pathlib, json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Optional
import openai
from dataclasses import dataclass

# Configure OpenAI (you'll need to set your API key)
# openai.api_key = "your-api-key-here"  # Or use environment variable

ROOT = pathlib.Path('.')
DEV_LOG = ROOT / '2026 SPRING FALL Development Sample Log_sensitive info removed.xlsx'
TEMPLATE = ROOT / 'Spec Template (1).xlsx'
BOM = ROOT / 'WORKING SIMPLIFIED BOM (1).xlsx'
OUTFILE = ROOT / 'AI-ENHANCED SPEC SHEETS.xlsx'

@dataclass
class MaterialMatch:
    original: str
    standardized: str
    confidence: float
    method: str  # 'exact', 'fuzzy', 'ai'

class AIEnhancedMaterialMatcher:
    def __init__(self, bom_materials: List[str], use_ai: bool = True):
        self.bom_materials = bom_materials
        self.use_ai = use_ai
        self.exact_lookup = {m.lower(): m for m in bom_materials if m}
        self.all_materials_lower = list(self.exact_lookup.keys())
        
        # Cache for AI results to avoid repeated API calls
        self.ai_cache = {}
        
    def standardize_material(self, raw_material: str) -> MaterialMatch:
        """Enhanced material standardization with AI fallback"""
        if not isinstance(raw_material, str) or not raw_material.strip():
            return MaterialMatch(raw_material, raw_material, 0.0, 'none')
        
        clean_material = raw_material.strip()
        material_lower = clean_material.lower()
        
        # 1. Try exact match first
        if material_lower in self.exact_lookup:
            return MaterialMatch(
                original=raw_material,
                standardized=self.exact_lookup[material_lower],
                confidence=1.0,
                method='exact'
            )
        
        # 2. Try fuzzy matching
        fuzzy_matches = difflib.get_close_matches(
            material_lower, self.all_materials_lower, n=1, cutoff=0.85
        )
        if fuzzy_matches:
            return MaterialMatch(
                original=raw_material,
                standardized=self.exact_lookup[fuzzy_matches[0]],
                confidence=0.85,
                method='fuzzy'
            )
        
        # 3. AI-enhanced matching (if enabled)
        if self.use_ai:
            ai_result = self._ai_material_match(clean_material)
            if ai_result:
                return ai_result
        
        # 4. No match found
        return MaterialMatch(raw_material, raw_material, 0.0, 'none')
    
    def _ai_material_match(self, material: str) -> Optional[MaterialMatch]:
        """Use AI to find the best material match"""
        # Check cache first
        if material in self.ai_cache:
            return self.ai_cache[material]
        
        try:
            # Create a prompt with our BOM materials
            bom_list = "\n".join([f"- {mat}" for mat in self.bom_materials[:50]])  # Limit for token efficiency
            
            prompt = f"""
You are a material standardization expert for footwear manufacturing.

Given this supplier material description: "{material}"

Find the best match from our standardized BOM materials:
{bom_list}

Return ONLY a JSON response with:
{{
    "best_match": "exact material name from BOM or null if no good match",
    "confidence": 0.0-1.0,
    "reasoning": "brief explanation"
}}

Requirements:
- Only return materials that exist exactly in the BOM list
- Confidence should be 0.7+ for a match
- Consider synonyms, abbreviations, and common material variations
- If no good match (confidence < 0.7), return null for best_match
"""

            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=200
            )
            
            result_text = response.choices[0].message.content.strip()
            result = json.loads(result_text)
            
            if result.get('best_match') and result.get('confidence', 0) >= 0.7:
                match = MaterialMatch(
                    original=material,
                    standardized=result['best_match'],
                    confidence=result['confidence'],
                    method='ai'
                )
                self.ai_cache[material] = match
                return match
            
        except Exception as e:
            print(f"⚠️  AI matching error for '{material}': {e}")
        
        return None

class AITextParser:
    """AI-enhanced parser for complex material descriptions"""
    
    def __init__(self, use_ai: bool = True):
        self.use_ai = use_ai
        # Fallback regex (same as before)
        self.tag_rx = re.compile(r'(^[A-Za-z/ &()]+:)', flags=re.M)
    
    def parse_material_block(self, text: str) -> Dict[str, str]:
        """Parse material descriptions with AI enhancement"""
        if not isinstance(text, str):
            return {}
        
        # Try AI parsing first
        if self.use_ai:
            ai_result = self._ai_parse_materials(text)
            if ai_result:
                return ai_result
        
        # Fallback to regex parsing
        return self._regex_parse_materials(text)
    
    def _ai_parse_materials(self, text: str) -> Optional[Dict[str, str]]:
        """Use AI to parse complex material descriptions"""
        try:
            prompt = f"""
Extract shoe part materials from this text: "{text}"

Common shoe parts include: Upper, Lining, Insole, Outsole, Midsole, Footbed, Heel, Toe, Quarter, Vamp, etc.

Return ONLY valid JSON with part names as keys and materials as values:
{{
    "Upper": "material name",
    "Lining": "material name",
    ...
}}

Rules:
- Only include parts that are clearly mentioned
- Clean up material names (remove extra spaces, fix obvious typos)
- Use standard part names (Upper not "upper part")
- If text is unclear or empty, return {{}}
"""

            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=300
            )
            
            result_text = response.choices[0].message.content.strip()
            return json.loads(result_text)
            
        except Exception as e:
            print(f"⚠️  AI parsing error: {e}")
            return None
    
    def _regex_parse_materials(self, text: str) -> Dict[str, str]:
        """Fallback regex parsing (original method)"""
        splits = self.tag_rx.split(text)
        parts = {}
        for tag, value in zip(splits[1::2], splits[2::2]):
            key = tag.replace(':', '').strip()
            parts[key] = value.replace('\n', ' ').strip()
        return parts

# Main enhanced processing function
def process_with_ai_enhancement(use_ai: bool = True):
    """Main function with AI enhancements"""
    
    print(f"🤖 AI Enhanced Spec Sheet Generator (AI: {'ON' if use_ai else 'OFF'})")
    print("=" * 60)
    
    # Load data (same as before)
    print("📋 Loading Development Sample Log...")
    dev = pd.read_excel(DEV_LOG)
    dev.columns = dev.iloc[0]
    dev = dev.drop(index=0).reset_index(drop=True)
    
    print("🔍 Loading Simplified BOM...")
    raw_bom = pd.read_excel(BOM, header=None).dropna(how='all').rename(columns={0:'raw'})
    raw_bom[['Part', 'Material']] = raw_bom['raw'].str.split(':', n=1, expand=True)
    raw_bom['Material'] = raw_bom['Material'].fillna('').str.strip()
    bom_materials = [m for m in raw_bom['Material'] if m]
    
    # Initialize AI components
    material_matcher = AIEnhancedMaterialMatcher(bom_materials, use_ai=use_ai)
    text_parser = AITextParser(use_ai=use_ai)
    
    print(f"📝 Loading template...")
    wb_master = load_workbook(TEMPLATE)
    base_sheet = wb_master.active
    
    # Process each sample
    processing_stats = {
        'total': 0,
        'successful': 0,
        'ai_matches': 0,
        'fuzzy_matches': 0,
        'exact_matches': 0,
        'no_matches': 0
    }
    
    for idx, row in dev.iterrows():
        try:
            sample_name = row['Sample Name']
            print(f"\n🔧 Processing: {sample_name}")
            
            # Create new sheet
            sheet = wb_master.copy_worksheet(base_sheet)
            clean_name = re.sub(r'[\\/*?[\]:]+', '_', str(sample_name))
            sheet.title = (clean_name[:28] + '...') if len(clean_name) > 31 else clean_name
            
            # Fill metadata
            season_gender = f"{row.get('Season', '')}, {row.get('Gender', '')}"
            sheet['B1'] = season_gender
            sheet['A4'] = row.get('Factory Ref #', '')
            sheet['C2'] = sample_name
            sheet['E2'] = row.get('Sample Order No.', '')
            
            # Enhanced material parsing
            upper_materials = text_parser.parse_material_block(row.get('Upper', ''))
            sole_materials = text_parser.parse_material_block(row.get('Sole', ''))
            
            # Enhanced material matching
            all_materials = {**upper_materials, **sole_materials}
            matched_materials = {}
            
            for part, raw_material in all_materials.items():
                match_result = material_matcher.standardize_material(raw_material)
                matched_materials[part] = match_result
                
                # Update stats
                if match_result.method == 'exact':
                    processing_stats['exact_matches'] += 1
                elif match_result.method == 'fuzzy':
                    processing_stats['fuzzy_matches'] += 1
                elif match_result.method == 'ai':
                    processing_stats['ai_matches'] += 1
                else:
                    processing_stats['no_matches'] += 1
                
                print(f"  {part}: {raw_material} → {match_result.standardized} ({match_result.method}, {match_result.confidence:.2f})")
            
            # Fill template
            def find_row(label: str) -> int:
                for r in range(1, sheet.max_row+1):
                    cell_value = sheet[f'A{r}'].value
                    if cell_value and str(cell_value).strip().startswith(label):
                        return r
                raise ValueError(f"Label {label} not found")
            
            materials_filled = 0
            for part, match_result in matched_materials.items():
                try:
                    r = find_row(f'{part}:')
                    sheet[f'B{r}'] = match_result.standardized
                    materials_filled += 1
                except ValueError:
                    pass
            
            processing_stats['successful'] += 1
            processing_stats['total'] += 1
            
        except Exception as e:
            print(f"  ❌ Error: {e}")
            processing_stats['total'] += 1
    
    # Save and report
    wb_master.remove(base_sheet)
    wb_master.save(OUTFILE)
    
    print(f"\n📊 PROCESSING SUMMARY")
    print(f"{'='*40}")
    print(f"Total samples: {processing_stats['total']}")
    print(f"Successful: {processing_stats['successful']}")
    print(f"Exact matches: {processing_stats['exact_matches']}")
    print(f"Fuzzy matches: {processing_stats['fuzzy_matches']}")
    print(f"AI matches: {processing_stats['ai_matches']}")
    print(f"No matches: {processing_stats['no_matches']}")
    print(f"\n✅ Created {OUTFILE}")

if __name__ == "__main__":
    # You can run with or without AI
    USE_AI = True  # Set to False to disable AI features
    process_with_ai_enhancement(use_ai=USE_AI) 