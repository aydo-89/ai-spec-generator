import re, difflib, pathlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = pathlib.Path('.')                                # folder with your files
DEV_LOG = ROOT / '2026 SPRING FALL Development Sample Log_sensitive info removed.xlsx'
TEMPLATE = ROOT / 'Spec Template (1).xlsx'
BOM      = ROOT / 'WORKING SIMPLIFIED BOM (1).xlsx'
OUTFILE  = ROOT / 'AUTO‑GENERATED SPEC SHEETS.xlsx'

# --------------------------------------------------
# 1.  Load & tidy the Development Sample Log
# --------------------------------------------------
print("📋 Loading Development Sample Log...")
dev = pd.read_excel(DEV_LOG)
print(f"Initial shape: {dev.shape}")
print(f"Initial columns: {list(dev.columns)}")

dev.columns = dev.iloc[0]                      # promote first row → header
dev = dev.drop(index=0).reset_index(drop=True) # drop the header row

print(f"After header promotion: {dev.shape}")
print(f"Final columns: {list(dev.columns)}")
print(f"Sample names: {dev['Sample Name'].tolist() if 'Sample Name' in dev.columns else 'Sample Name column not found'}")

# --------------------------------------------------
# 2.  Build a material normaliser from the Simplified BOM
# --------------------------------------------------
print("\n🔍 Loading Simplified BOM...")
raw_bom = (pd.read_excel(BOM, header=None)
             .dropna(how='all')                # remove blank rows
             .rename(columns={0:'raw'}) )

print(f"BOM shape: {raw_bom.shape}")
print(f"First few BOM entries: {raw_bom['raw'].head().tolist()}")

# split "Part: Material name"
raw_bom[['Part', 'Material']] = raw_bom['raw'].str.split(':', n=1, expand=True)
raw_bom['Part'] = raw_bom['Part'].str.strip()
raw_bom['Material'] = raw_bom['Material'].fillna('').str.strip()

# dictionaries for exact & fuzzy look‑up
exact_lookup = {m.lower(): m for m in raw_bom['Material'] if m}
all_materials = list(exact_lookup.keys())

print(f"Loaded {len(exact_lookup)} materials for lookup")

def standardise(text: str) -> str:
    """Return canonical material name from the BOM (exact → fuzzy)."""
    if not isinstance(text, str):
        return text
    txt = text.lower()
    if txt in exact_lookup:
        return exact_lookup[txt]

    # fuzzy fallback
    match = difflib.get_close_matches(txt, all_materials, n=1, cutoff=0.80)
    return exact_lookup[match[0]] if match else text   # leave unchanged if no hit

# --------------------------------------------------
# 3.  Helper to parse the long 'Upper / Sole …' cells
# --------------------------------------------------
TAG_RX = re.compile(r'(^[A-Za-z/ &()]+:)', flags=re.M)   # find "Upper:", "Lining:" etc.

def explode_parts(long_cell: str) -> dict:
    """
    Turn a block like:
    'Upper:\nLeather ABC\nLining: Microfiber\nSole: …'
    …into {'Upper':'Leather ABC', 'Lining':'Microfiber', …}
    """
    if not isinstance(long_cell, str):
        return {}
    splits = TAG_RX.split(long_cell)
    parts = {}
    for tag, value in zip(splits[1::2], splits[2::2]):
        key = tag.replace(':','').strip()
        parts[key] = standardise(value.replace('\n',' ').strip())
    return parts

# --------------------------------------------------
# 4.  Open the blank template once; then duplicate & fill
# --------------------------------------------------
print(f"\n📝 Loading template from {TEMPLATE}...")
wb_master = load_workbook(TEMPLATE)
base_sheet: Worksheet = wb_master.active      # 'Spec Template' sheet

print(f"Template loaded. Active sheet: {base_sheet.title}")
print(f"Template dimensions: {base_sheet.max_row} rows x {base_sheet.max_column} columns")

processed_count = 0
for idx, row in dev.iterrows():
    try:
        sample_name = row['Sample Name']
        print(f"\n🔧 Processing sample: {sample_name}")
        
        sheet = wb_master.copy_worksheet(base_sheet)      # duplicate template
        # Ensure sheet name is valid (Excel sheet names can't exceed 31 chars and have certain restrictions)
        clean_name = re.sub(r'[\\/*?[\]:]+', '_', str(sample_name))  # Replace invalid characters
        sheet.title = (clean_name[:28] + '...') if len(clean_name) > 31 else clean_name

        # 4‑A  Top‑of‑sheet metadata
        season_gender = f"{row.get('Season', '')}, {row.get('Gender', '')}"
        sheet['B1'] = season_gender                       # "Season Created:" cell
        sheet['A4'] = row.get('Factory Ref #', '')       # factory ref
        sheet['C2'] = sample_name                         # sample name
        sheet['E2'] = row.get('Sample Order No.', '')    # sample order #

        # 4‑B  Material sections
        upper_dict = explode_parts(row.get('Upper', ''))
        sole_dict  = explode_parts(row.get('Sole', ''))

        print(f"  Upper materials: {upper_dict}")
        print(f"  Sole materials: {sole_dict}")

        def find_row(label: str) -> int:
            """Locate the row containing a given label in column A."""
            for r in range(1, sheet.max_row+1):
                cell_value = sheet[f'A{r}'].value
                if cell_value and str(cell_value).strip().startswith(label):
                    return r
            raise ValueError(f"Label {label} not found")

        # Fill in materials
        materials_filled = 0
        for part, material in {**upper_dict, **sole_dict}.items():
            try:
                r = find_row(f'{part}:')
                sheet[f'B{r}'] = material
                materials_filled += 1
                print(f"  └─ {part}: {material} → Row {r}")
            except ValueError as e:
                print(f"  ⚠️  {e} - skipping {part}")

        print(f"  ✅ Filled {materials_filled} materials")
        processed_count += 1
        
    except Exception as e:
        print(f"  ❌ Error processing {row.get('Sample Name', 'Unknown')}: {e}")
        continue

# --------------------------------------------------
# 5.  Save
# --------------------------------------------------
print(f"\n💾 Saving workbook...")
wb_master.remove(base_sheet)        # drop the unused original
wb_master.save(OUTFILE)
print(f'✅ Created {OUTFILE}')
print(f'📊 Successfully processed {processed_count} samples') 